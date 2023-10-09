# Import packages

import openpyxl
from openpyxl.styles import PatternFill
from selenium import webdriver
import time
import requests
import datetime
import os
import XLUtils

# Initalising webdriver
driver = webdriver.Chrome(executable_path="D:\chromedriver_win32\chromedriver.exe")

#Title of the page
driver.get("https://chartink.com/screener/stocks-near-52-week-high-down-by-20-3")
time.sleep(2)

#to Maximize the window
driver.maximize_window()

#To print the title of the page
print("Title: " + driver.title)
time.sleep(1)

#To Scroll the page
driver.execute_script("window.scrollTo(0,500)","")
time.sleep(2)

#Run the scan
button=driver.find_elements_by_xpath("//*[@id=\'root\']/div[2]/div/div/div/div[2]/div/div[2]/div[2]/div[1]/button[1]")
button[0].click()
time.sleep(2)

#Download the Excel
buttonExcel=driver.find_elements_by_xpath("//*[@id=\'DataTables_Table_0_wrapper\']/div[1]/div/button[3]")
buttonExcel[0].click()
time.sleep(2)

#
# wb = openpyxl.load_workbook("DMT Tech  Funda Scan, Technical Analysis Scanner.xlsx")
# ws = wb['Sheet1']
# red = PatternFill(patternType='solid', fgColor='EE1515')
# blue = PatternFill(patternType='solid', fgColor='00CCFF')
# green = PatternFill(patternType='solid', fgColor='09A834')
# yellow = PatternFill(patternType='solid', fgColor='EEEE0D')

#initializing the variable to store the date and timestamp
today = datetime.datetime.now()
now = today.strftime("%Y-%m-%d %H:%M:%S")

#Creating a variable to store path of Excel file
path_of_ExcelFile="C:\\Users\\LENOVO\\Downloads\\DMT Tech  Funda Scan, Technical Analysis Scanner.xlsx"

#Initialising the dictionary to store excel path
files = {'document':open(path_of_ExcelFile,'rb')}

#Write the TradingView link in the Excel File
rows=XLUtils.getRowCount(path_of_ExcelFile,'Sheet1')
print(rows)

XLUtils.writeData(path_of_ExcelFile, "Sheet1", 2, 8, "Tradingview Link")
XLUtils.writeData(path_of_ExcelFile, "Sheet1", 2, 9, "Tradingview HyperLink")
XLUtils.writeData(path_of_ExcelFile, "Sheet1", 2, 10, "Symbol,")

for r in range(3, rows + 1):
    symbol = XLUtils.readData(path_of_ExcelFile, "Sheet1", r, 3)
    XLUtils.writeData(path_of_ExcelFile, "Sheet1", r, 8, "https://in.tradingview.com/chart/THCdg7nW/?symbol="+symbol)
    XLUtils.writeData(path_of_ExcelFile, "Sheet1", r, 9, "=hyperlink(H8,C" + str(r) + ")")
    XLUtils.writeData(path_of_ExcelFile, "Sheet1", r, 10, symbol + ",")

#Posting the Excel on Telegram Channel
#Put bot ID and chat ID
resp = requests.post('https://api.telegram.org/bot-ID/sendDocument?chat_id=chatID&caption= {}'.format(now), files=files)

#Print the response status code
print(resp.status_code)
time.sleep(2)

#Storing the file from dictionary into excel
excel=files['document']
excel.close()

#Close the browser
driver.close()

#Delete the Excel file present in the PC
os.remove(path_of_ExcelFile)

